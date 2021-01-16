import xml.etree.ElementTree as ET
import openpyxl
from openpyxl import Workbook

path = "C:\\Users\\rafwz\\Documents\\PythonProjects\\tutorial\\Verdictsheet_Maserati_09.00.00.249.xlsx"
i = 0

def innertext(tag):
  return (tag.text or '') + ''.join(innertext(e) for e in tag) + (tag.tail or '')

#def search_word(word, excel_file):


root = ET.parse('Report_DIAG_ALL2.xml').getroot()

test_result = []
tc_names = []

for testcase in root.iter('testcase'):
    name = testcase.find('title')
    #print(innertext(name))
    tc_names.append(innertext(name))
    result = testcase.find('verdict').attrib['result']
    test_result.append([innertext(name), result])

#print(test_result[0][0])

#wb_obj.save();

#wb_obj = Workbook("Workitem_VCAN.xlsx")
#sheets = wb_obj.sheetnames
#Sheet1 = wb_obj[sheets[1]]
#print(wb_obj.sheetnames)
#Sheet1.cell(row=1, column=1).value = "test"
#wb_obj.save()

#from openpyxl import Workbook
#import openpyxl

#file = "enter_path_to_file_here"
#wb = openpyxl.load_workbook(file, read_only=True)
#ws = wb.active

#for row in ws.iter_rows("E"):
#    for cell in row:
#        if cell.value == "ABC":
#            print(ws.cell(row=cell.row, column=2).value) #change column number for any cell value you want
#ws = wb_obj.active

wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.sheetnames
workbook = wb_obj.active
ws = wb_obj["DiagComMgr"]
Sheet1 = wb_obj[sheet_obj[1]]


#for work_cell in test_result:
#    Sheet1.cell(row=9+i, column=4).value = work_cell[0]
#    Sheet1.cell(row=9+i, column=10).value = work_cell[1]
#    i= i+1

for row in ws.iter_rows(min_col=2, max_col=2, min_row = 9):
    for cell in row:
        for req in tc_names:
            #print(req)
            #print(cell.value)
            req_present = req in str(cell.value)
            if req_present:
                print("dupa")
                #ws.cell(row=cell.row, column=10).value = test_result[tc_names.index(req)][1]
                ws.cell(row=cell.row, column=11).value = "dip"
#for row in list(Sheet1.rows[1:]):
#    if row[1].value.find("ECU") != -1:
#        location = row

#print(location)
wb_obj.save("test.xlsx")

